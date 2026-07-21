# -*- coding: utf-8 -*-
"""
Construye/actualiza la base de datos de flota EDINSA: fleet_db.json

Consolida en un solo archivo (que se va actualizando mes a mes, sin perder
histórico) dos fuentes:

  1. "BD PLACAS EDINSA.xlsx" (hoja "Hoja1")
     Columnas: Placa | Marca | Linea | Modelo | Plaza
     -> de aquí sale la clasificación de cada vehículo por Plaza (T1-CENTRO,
        T1-OCCIDENTE, T1-COSTA, T1-SANTANDERES, T1-ANTIOQUIA, etc.)

  2. "data - *.xlsx" (hoja "Export") — el más reciente por fecha de modificación
     Columna 1 = Placa, columnas siguientes = un mes cada una (fecha en el
     encabezado) con los km recorridos ese mes.

Las claves de vehículo en ambas fuentes son la PLACA pura, sin prefijo
"CO_" (ese prefijo solo aparece en el campo "Nombre de Máquina" de los CSV
de eventos de cámaras — ver build_data.py, que ya lo quita al leer).

Merge incremental: si fleet_db.json ya existe, se parte de él y se
sobreescriben/agregan placas y meses de km con lo que traigan los archivos
de origen actuales. Así, cada mes puedes reemplazar "BD PLACAS EDINSA.xlsx"
y/o soltar un nuevo "data - <fecha>.xlsx" con el mes más reciente y volver a
correr este script: lo histórico ya guardado no se pierde.

Uso:
    python build_fleet_db.py
"""
import glob
import json
import os
import sys
from datetime import datetime

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, "fleet_db.json")
PLACAS_FILE = os.path.join(BASE_DIR, "BD PLACAS EDINSA.xlsx")


def find_latest_km_file():
    """Encuentra el archivo de export de km más reciente: 'data - *.xlsx'
    (ignora archivos de bloqueo de Excel que empiezan con '~$')."""
    candidates = [
        f for f in glob.glob(os.path.join(BASE_DIR, "data*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    ]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def load_db():
    if os.path.exists(DB_FILE):
        with open(DB_FILE, encoding="utf-8") as fh:
            return json.load(fh)
    return {"actualizado": None, "placas": {}, "km_mensual": {}, "fuentes": {}}


def import_placas(db, path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Hoja1"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    n = 0
    for row in rows[1:]:
        if not row or not row[idx.get("Placa", 0)]:
            continue
        placa = str(row[idx["Placa"]]).strip().upper()
        db["placas"][placa] = {
            "marca": row[idx.get("Marca", 1)],
            "linea": row[idx.get("Linea", 2)],
            "modelo": row[idx.get("Modelo", 3)],
            "plaza": row[idx.get("Plaza", 4)],
        }
        n += 1
    print(f"  Placas importadas/actualizadas: {n}")


def import_km(db, path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Export"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    month_cols = []  # (col_index, "YYYY-MM")
    for i, h in enumerate(header):
        if i == 0:
            continue
        if isinstance(h, datetime):
            month_cols.append((i, h.strftime("%Y-%m")))
    n_placas = 0
    n_celdas = 0
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        placa = str(row[0]).strip().upper()
        bucket = db["km_mensual"].setdefault(placa, {})
        n_placas += 1
        for col_i, mkey in month_cols:
            v = row[col_i] if col_i < len(row) else None
            if v is None:
                continue
            bucket[mkey] = round(float(v), 2)
            n_celdas += 1
    print(f"  Vehículos con km importados/actualizados: {n_placas} ({n_celdas} celdas mes-vehículo)")


def main():
    db = load_db()

    if os.path.exists(PLACAS_FILE):
        print(f"Leyendo {os.path.basename(PLACAS_FILE)}…")
        import_placas(db, PLACAS_FILE)
        db["fuentes"]["placas"] = os.path.basename(PLACAS_FILE)
    else:
        print(f"[WARN] No se encontró {PLACAS_FILE}; se mantiene lo que ya había en la BD.")

    km_file = find_latest_km_file()
    if km_file:
        print(f"Leyendo {os.path.basename(km_file)}…")
        import_km(db, km_file)
        db["fuentes"]["km"] = os.path.basename(km_file)
    else:
        print("[WARN] No se encontró ningún archivo 'data*.xlsx' con km; se mantiene lo que ya había en la BD.")

    db["actualizado"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    with open(DB_FILE, "w", encoding="utf-8") as fh:
        json.dump(db, fh, ensure_ascii=False, separators=(",", ":"))

    todos_los_meses = sorted({m for bucket in db["km_mensual"].values() for m in bucket})
    plazas = sorted({p["plaza"] for p in db["placas"].values() if p.get("plaza")})
    print(f"\n[OK] {DB_FILE}")
    print(f"  Placas totales en BD: {len(db['placas'])}")
    print(f"  Vehículos con histórico de km: {len(db['km_mensual'])}")
    print(f"  Meses de km disponibles: {todos_los_meses[0] if todos_los_meses else '—'} -> {todos_los_meses[-1] if todos_los_meses else '—'} ({len(todos_los_meses)} meses)")
    print(f"  Plazas: {plazas}")


if __name__ == "__main__":
    main()
